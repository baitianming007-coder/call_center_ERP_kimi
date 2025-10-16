"""
工具函数模块
"""
from datetime import datetime
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_to_excel(data, headers, filename='export.xlsx'):
    """
    导出数据到 Excel
    
    Args:
        data: 数据列表（每行是一个字典或列表）
        headers: 表头列表
        filename: 文件名
        
    Returns:
        BytesIO: Excel 文件的字节流
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl 未安装，无法导出 Excel")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    
    # 样式设置
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="E5E7EB")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        if isinstance(row_data, dict):
            values = [row_data.get(key, '') for key in headers]
        else:
            values = row_data
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def format_date(date_str, format='%Y-%m-%d'):
    """格式化日期字符串"""
    if not date_str:
        return ''
    try:
        if isinstance(date_str, str):
            date_obj = datetime.strptime(date_str, format)
            return date_obj.strftime('%Y年%m月%d日')
        return date_str
    except:
        return date_str


def validate_phone(phone):
    """验证手机号格式"""
    if not phone:
        return False
    return len(phone) == 11 and phone.isdigit() and phone.startswith('1')


def format_currency(amount):
    """格式化货币"""
    try:
        return f"¥{float(amount):,.2f}"
    except:
        return "¥0.00"



