"""
数据导出模块 - Excel & PDF
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import Image
try:
    from PyPDF2 import PdfReader, PdfWriter
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False


def export_employees_to_excel(employees):
    """
    导出员工数据到 Excel
    
    Args:
        employees: 员工列表
        
    Returns:
        BytesIO: Excel 文件字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "员工数据"
    
    # 样式设置
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="E5E7EB")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # 表头
    headers = ["工号", "姓名", "手机号", "团队", "状态", "入职日期", "在职状态"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据
    status_map = {
        'trainee': '培训期',
        'C': 'C级',
        'B': 'B级',
        'A': 'A级',
        'eliminated': '已淘汰'
    }
    
    for row_idx, emp in enumerate(employees, 2):
        data = [
            emp['employee_no'],
            emp['name'],
            emp['phone'] or '-',
            emp['team'],
            status_map.get(emp['status'], emp['status']),
            emp['join_date'],
            '在职' if emp['is_active'] else '离职'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_performance_to_excel(performance_list):
    """导出业绩数据到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "业绩数据"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="E5E7EB")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # 表头
    headers = ["日期", "工号", "姓名", "团队", "状态", "出单数", "日提成", "有效工作日"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据
    status_map = {'trainee': '培训期', 'C': 'C级', 'B': 'B级', 'A': 'A级', 'eliminated': '已淘汰'}
    
    for row_idx, perf in enumerate(performance_list, 2):
        data = [
            perf['work_date'],
            perf['employee_no'],
            perf['name'],
            perf['team'],
            status_map.get(perf['status'], perf['status']),
            perf['orders_count'],
            perf['commission'],
            '是' if perf['is_valid_workday'] else '否'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            
            # 数值类型右对齐
            if col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_salary_to_excel(salary_list, year_month):
    """导出薪资数据到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year_month}薪资"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="E5E7EB")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # 标题行
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"{year_month} 月度薪资统计表"
    title_cell.font = Font(bold=True, size=16, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["工号", "姓名", "团队", "状态", "底薪/固定", "全勤奖", "绩效奖", "提成", "总计"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据
    status_map = {'trainee': '培训期', 'C': 'C级', 'B': 'B级', 'A': 'A级', 'eliminated': '已淘汰'}
    total_base = 0
    total_attendance = 0
    total_performance = 0
    total_commission = 0
    total_salary = 0
    
    for row_idx, sal in enumerate(salary_list, 3):
        data = [
            sal['employee_no'],
            sal['name'],
            sal['team'],
            status_map.get(sal['status'], sal['status']),
            sal['base_salary'],
            sal['attendance_bonus'],
            sal['performance_bonus'],
            sal['commission'],
            sal['total_salary']
        ]
        
        total_base += sal['base_salary']
        total_attendance += sal['attendance_bonus']
        total_performance += sal['performance_bonus']
        total_commission += sal['commission']
        total_salary += sal['total_salary']
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            
            # 数值列右对齐
            if col_idx >= 5:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # 数值格式
                if col_idx >= 5:
                    cell.number_format = '¥#,##0.00'
    
    # 合计行
    summary_row = len(salary_list) + 3
    summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    
    ws.cell(row=summary_row, column=1, value="合计")
    for col_idx in range(1, 10):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col_idx <= 4 else "right", vertical="center")
    
    ws.cell(row=summary_row, column=5, value=total_base).number_format = '¥#,##0.00'
    ws.cell(row=summary_row, column=6, value=total_attendance).number_format = '¥#,##0.00'
    ws.cell(row=summary_row, column=7, value=total_performance).number_format = '¥#,##0.00'
    ws.cell(row=summary_row, column=8, value=total_commission).number_format = '¥#,##0.00'
    ws.cell(row=summary_row, column=9, value=total_salary).number_format = '¥#,##0.00'
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_salary_pdf(employee_data, salary_data, year_month, password=None, add_watermark=False):
    """
    生成个人薪资单 PDF (P2-10增强版)
    
    Args:
        employee_data: 员工信息
        salary_data: 薪资数据
        year_month: 年月
        password: PDF密码保护（可选）
        add_watermark: 是否添加水印
        
    Returns:
        BytesIO: PDF 文件字节流
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        topMargin=60,
        bottomMargin=60,
        leftMargin=50,
        rightMargin=50
    )
    
    # 构建文档内容
    story = []
    styles = getSampleStyleSheet()
    
    # 公司Logo占位符（将来可替换）
    header_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6B7280'),
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    company_header = Paragraph("📞 呼叫中心职场管理系统", header_style)
    story.append(company_header)
    
    divider_style = ParagraphStyle(
        'Divider',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#D1D5DB'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    story.append(Paragraph("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", divider_style))
    
    # 标题样式（增强）
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # 添加标题
    title = Paragraph("月度薪资单", title_style)
    story.append(title)
    
    subtitle = Paragraph(f"{year_month}", subtitle_style)
    story.append(subtitle)
    story.append(Spacer(1, 10))
    
    # 员工信息表
    employee_info = [
        ['工号：', employee_data['employee_no'], '姓名：', employee_data['name']],
        ['团队：', employee_data['team'], '状态：', employee_data['status']],
    ]
    
    info_table = Table(employee_info, colWidths=[1*inch, 2*inch, 1*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6B7280')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#6B7280')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#111827')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # 薪资明细表
    salary_data_table = [
        ['项目', '金额'],
        ['底薪/固定薪资', f"¥{salary_data['base_salary']:.2f}"],
        ['全勤奖', f"¥{salary_data['attendance_bonus']:.2f}"],
        ['绩效奖', f"¥{salary_data['performance_bonus']:.2f}"],
        ['日提成', f"¥{salary_data['commission']:.2f}"],
        ['', ''],
        ['总计', f"¥{salary_data['total_salary']:.2f}"],
    ]
    
    salary_table = Table(salary_data_table, colWidths=[3*inch, 3*inch])
    salary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#F3F4F6')),
        ('FONTSIZE', (0, 6), (-1, 6), 14),
        ('TEXTCOLOR', (1, 6), (1, 6), colors.HexColor('#2563EB')),
        ('GRID', (0, 0), (-1, 5), 1, colors.HexColor('#E5E7EB')),
        ('LINEABOVE', (0, 6), (-1, 6), 2, colors.HexColor('#2563EB')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    story.append(salary_table)
    story.append(Spacer(1, 30))
    
    # 计算明细
    detail_style = ParagraphStyle(
        'Detail',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        leading=16
    )
    
    story.append(Paragraph('<b>计算明细：</b>', detail_style))
    story.append(Spacer(1, 10))
    
    detail_text = salary_data['calculation_detail'].replace('\n', '<br/>')
    story.append(Paragraph(detail_text, detail_style))
    
    story.append(Spacer(1, 40))
    
    # 页脚
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=TA_CENTER
    )
    
    footer_text = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>呼叫中心职场管理系统"
    story.append(Paragraph(footer_text, footer_style))
    
    # 如果设置了水印
    if add_watermark:
        watermark_style = ParagraphStyle(
            'Watermark',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#E5E7EB'),
            alignment=TA_CENTER
        )
        story.append(Spacer(1, 20))
        story.append(Paragraph("本薪资单仅供个人查阅，请勿外传", watermark_style))
    
    # 生成 PDF
    doc.build(story)
    buffer.seek(0)
    
    # P2-10: 添加密码保护
    if password and HAS_PYPDF2:
        try:
            # 读取刚生成的PDF
            pdf_reader = PdfReader(buffer)
            pdf_writer = PdfWriter()
            
            # 复制所有页面
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)
            
            # 设置密码保护
            pdf_writer.encrypt(user_password=password, owner_password=password, use_128bit=True)
            
            # 写入新的buffer
            protected_buffer = io.BytesIO()
            pdf_writer.write(protected_buffer)
            protected_buffer.seek(0)
            
            return protected_buffer
        except Exception as e:
            # 如果加密失败，返回未加密的PDF
            print(f"PDF加密失败: {str(e)}")
            buffer.seek(0)
            return buffer
    elif password and not HAS_PYPDF2:
        # 提示需要安装PyPDF2
        print("警告：未安装PyPDF2，无法设置PDF密码保护")
    
    return buffer



