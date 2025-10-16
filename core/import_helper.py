#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据导入辅助模块
支持Excel批量导入业绩数据
"""

import openpyxl
from datetime import datetime
from core.database import get_db, query_db
import re

class ImportValidator:
    """导入数据验证器"""
    
    @staticmethod
    def validate_employee_no(employee_no):
        """验证工号格式"""
        if not employee_no:
            return False, "工号不能为空"
        
        # 检查工号是否存在
        employee = query_db(
            'SELECT id FROM employees WHERE employee_no = ? AND is_active = 1',
            (employee_no,),
            one=True
        )
        
        if not employee:
            return False, f"工号 {employee_no} 不存在或已离职"
        
        return True, employee['id']
    
    @staticmethod
    def validate_date(date_str):
        """验证日期格式"""
        if not date_str:
            return False, "日期不能为空"
        
        try:
            # 支持多种日期格式
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
                try:
                    date_obj = datetime.strptime(str(date_str), fmt)
                    return True, date_obj.strftime('%Y-%m-%d')
                except:
                    continue
            
            # 如果是datetime对象
            if isinstance(date_str, datetime):
                return True, date_str.strftime('%Y-%m-%d')
            
            return False, f"日期格式错误: {date_str}，应为 YYYY-MM-DD"
        except Exception as e:
            return False, f"日期解析失败: {str(e)}"
    
    @staticmethod
    def validate_orders(orders):
        """验证出单数"""
        if orders is None or orders == '':
            return False, "出单数不能为空"
        
        try:
            orders_int = int(orders)
            if orders_int < 0:
                return False, "出单数不能为负数"
            if orders_int > 100:
                return False, "出单数超过合理范围（最大100）"
            return True, orders_int
        except:
            return False, f"出单数必须是整数: {orders}"

class ExcelImporter:
    """Excel导入器"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.fail_count = 0
    
    def validate_template(self, workbook):
        """验证Excel模板格式"""
        sheet = workbook.active
        
        # 检查表头
        expected_headers = ['工号', '日期', '出单数']
        actual_headers = [cell.value for cell in sheet[1]]
        
        if actual_headers[:3] != expected_headers:
            return False, f"模板格式错误，表头应为: {expected_headers}，实际为: {actual_headers[:3]}"
        
        return True, "模板格式正确"
    
    def parse_excel(self, file_path):
        """解析Excel文件"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # 验证模板
            valid, message = self.validate_template(workbook)
            if not valid:
                return None, message
            
            sheet = workbook.active
            data_rows = []
            
            # 从第2行开始读取数据（跳过表头）
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row[:3]):  # 跳过空行
                    continue
                
                employee_no, date_str, orders = row[:3]
                
                data_rows.append({
                    'row_number': row_idx,
                    'employee_no': employee_no,
                    'date': date_str,
                    'orders': orders
                })
            
            return data_rows, None
        except Exception as e:
            return None, f"Excel文件解析失败: {str(e)}"
    
    def validate_data(self, data_rows):
        """验证数据"""
        validated_data = []
        
        for row in data_rows:
            row_num = row['row_number']
            errors = []
            
            # 验证工号
            valid, result = ImportValidator.validate_employee_no(row['employee_no'])
            if not valid:
                errors.append(f"第{row_num}行: {result}")
                continue
            employee_id = result
            
            # 验证日期
            valid, result = ImportValidator.validate_date(row['date'])
            if not valid:
                errors.append(f"第{row_num}行: {result}")
                continue
            date_str = result
            
            # 验证出单数
            valid, result = ImportValidator.validate_orders(row['orders'])
            if not valid:
                errors.append(f"第{row_num}行: {result}")
                continue
            orders = result
            
            if errors:
                self.errors.extend(errors)
                self.fail_count += 1
            else:
                validated_data.append({
                    'employee_id': employee_id,
                    'date': date_str,
                    'orders': orders,
                    'row_number': row_num
                })
        
        return validated_data
    
    def check_duplicates(self, validated_data):
        """检查重复记录"""
        from core.database import query_db
        
        for item in validated_data:
            existing = query_db(
                'SELECT id FROM performance WHERE employee_id = ? AND date = ?',
                (item['employee_id'], item['date']),
                one=True
            )
            
            if existing:
                self.warnings.append(
                    f"第{item['row_number']}行: 该员工在 {item['date']} 已有业绩记录，将更新数据"
                )
                item['is_update'] = True
            else:
                item['is_update'] = False
    
    def import_data(self, validated_data, user_id):
        """导入数据到数据库"""
        from core.database import get_db
        from core.salary_engine import calculate_daily_commission
        from core.audit import log_operation
        
        db = get_db()
        
        for item in validated_data:
            try:
                # 计算提成
                employee = query_db(
                    'SELECT status FROM employees WHERE id = ?',
                    (item['employee_id'],),
                    one=True
                )
                
                commission = calculate_daily_commission(
                    item['orders'],
                    employee['status']
                )
                
                if item['is_update']:
                    # 更新现有记录
                    db.execute(
                        '''UPDATE performance 
                           SET orders = ?, commission = ?, updated_at = CURRENT_TIMESTAMP
                           WHERE employee_id = ? AND date = ?''',
                        (item['orders'], commission, item['employee_id'], item['date'])
                    )
                    action = 'update'
                else:
                    # 插入新记录
                    db.execute(
                        '''INSERT INTO performance (employee_id, date, orders, commission)
                           VALUES (?, ?, ?, ?)''',
                        (item['employee_id'], item['date'], item['orders'], commission)
                    )
                    action = 'create'
                
                db.commit()
                
                # 记录日志
                log_operation(
                    user_id=user_id,
                    action=action,
                    target_type='performance',
                    target_id=item['employee_id'],
                    details=f"Excel导入: {item['date']} - {item['orders']}单"
                )
                
                self.success_count += 1
                
            except Exception as e:
                self.errors.append(f"第{item['row_number']}行导入失败: {str(e)}")
                self.fail_count += 1
                db.rollback()
    
    def import_from_file(self, file_path, user_id):
        """从Excel文件导入数据"""
        # 解析Excel
        data_rows, error = self.parse_excel(file_path)
        if error:
            return {
                'success': False,
                'message': error,
                'errors': [error]
            }
        
        if not data_rows:
            return {
                'success': False,
                'message': 'Excel文件中没有数据',
                'errors': ['文件为空']
            }
        
        # 验证数据
        validated_data = self.validate_data(data_rows)
        
        if not validated_data:
            return {
                'success': False,
                'message': '所有数据验证失败',
                'errors': self.errors,
                'fail_count': self.fail_count
            }
        
        # 检查重复
        self.check_duplicates(validated_data)
        
        # 导入数据
        self.import_data(validated_data, user_id)
        
        return {
            'success': True,
            'message': f'导入完成：成功 {self.success_count} 条，失败 {self.fail_count} 条',
            'success_count': self.success_count,
            'fail_count': self.fail_count,
            'errors': self.errors,
            'warnings': self.warnings
        }

def generate_import_template():
    """生成Excel导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "业绩导入模板"
    
    # 设置表头
    headers = ['工号', '日期', '出单数']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加示例数据（3行）
    today = datetime.now()
    examples = [
        ('a001', today.strftime('%Y-%m-%d'), 5),
        ('a002', (today - timedelta(days=1)).strftime('%Y-%m-%d'), 3),
        ('a003', (today - timedelta(days=2)).strftime('%Y-%m-%d'), 8),
    ]
    
    for row_idx, (emp_no, date, orders) in enumerate(examples, start=2):
        ws.cell(row=row_idx, column=1, value=emp_no)
        ws.cell(row=row_idx, column=2, value=date)
        ws.cell(row=row_idx, column=3, value=orders)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # 添加说明
    ws.cell(row=6, column=1, value="说明：")
    ws.cell(row=7, column=1, value="1. 工号必须存在且为在职员工")
    ws.cell(row=8, column=1, value="2. 日期格式：YYYY-MM-DD（如2025-01-15）")
    ws.cell(row=9, column=1, value="3. 出单数必须为0-100的整数")
    ws.cell(row=10, column=1, value="4. 如果日期已有记录，将更新数据")
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


